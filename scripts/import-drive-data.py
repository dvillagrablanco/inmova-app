#!/usr/bin/env python3
"""
Import Rovida/Viroda data from Google Drive Excel files into INMOVA DB.
Executed via SSH on the production server.
"""
import sys
import json
import time

sys.path.insert(0, '/home/ubuntu/.local/lib/python3.12/site-packages')
import paramiko
import openpyxl

SERVER = '157.180.119.236'
USER = 'root'
PASSWORD = 'hBXxC6pZCQPBLPiHGUHkASiln+Su/BAVQAN6qQ+xjVo='

def exec_node(client, script, timeout=30):
    cmd = f'cd /opt/inmova-app && node -e "{script}" 2>/dev/null'
    stdin, stdout, stderr = client.exec_command(cmd, timeout=timeout)
    exit_status = stdout.channel.recv_exit_status()
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    return out, err

def parse_rovida_inmuebles():
    wb = openpyxl.load_workbook('/tmp/INMUEBLES_ROVIDA.xlsx')
    buildings = []
    
    sheet_map = {
        'OFICINA AVD.EUROPA MADRID': {'address': 'Avd. Europa 34B esc izq, 1ºizq, Madrid', 'city': 'Madrid', 'type': 'oficina'},
        'LOCAL BARQUILLO MADRID': {'address': 'Barquillo 30, Madrid', 'city': 'Madrid', 'type': 'local'},
        'LOCALES REINA MADRID': {'address': 'Reina, Madrid', 'city': 'Madrid', 'type': 'local'},
        'EDIFICIO PIAMONTE 23 MADRID': {'address': 'C/Piamonte 23, Madrid', 'city': 'Madrid', 'type': 'oficina'},
        'NAVE AVD CUBA PALENCIA': {'address': 'Avd. Cuba, Palencia', 'city': 'Palencia', 'type': 'nave_industrial'},
        'OFICINA VALLADOLID': {'address': 'Constitución 8, 2º izq, Valladolid', 'city': 'Valladolid', 'type': 'oficina'},
        'OFICINA PALENCIA': {'address': 'Palencia', 'city': 'Palencia', 'type': 'oficina'},
        'PLAZAS ESPRONCEDA(MADRID)': {'address': 'C/Espronceda, Madrid', 'city': 'Madrid', 'type': 'garaje'},
        'PLAZAS VALLADOLID': {'address': 'Constitución 5, Valladolid', 'city': 'Valladolid', 'type': 'garaje'},
        'PLAZAS M.PELAYO(PALENCIA)': {'address': 'M. Pelayo, Palencia', 'city': 'Palencia', 'type': 'garaje'},
        'APARTAMENTOS BENIDORM': {'address': 'C/Ibiza, Benidorm', 'city': 'Benidorm', 'type': 'vivienda'},
    }
    
    for sheet_name in wb.sheetnames:
        info = sheet_map.get(sheet_name, {'address': sheet_name, 'city': 'Madrid', 'type': 'vivienda'})
        ws = wb[sheet_name]
        tenants = []
        
        for row in range(2, ws.max_row + 1):
            row_data = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
            inquilino = None
            renta = None
            
            for val in row_data:
                if val and isinstance(val, str) and len(val) > 3 and not val.startswith('=') and val not in ('INQUILINO', 'FINCA', 'Módulo', 'Nº PLAZA', 'PLANTA', 'Propietario', 'Nº HABITACIONES'):
                    if any(c.isalpha() for c in val) and not val.startswith('20') and val not in ('X', 'LIBRE', 'OCUPADA', 'USO EMPRESA', 'EN VENTA', 'EMPLEADO', 'VACIO', 'SÓTANO', 'en efectivo'):
                        if not inquilino:
                            inquilino = val.strip()
                for val2 in row_data:
                    if isinstance(val2, (int, float)) and val2 > 10:
                        renta = float(val2)
            
            if inquilino and inquilino not in ('RENTA SIN IVA', 'RENTA CON IVA', 'FECHA CONTRATO', 'FIN CONTRATO', 'FECHA INICIO CONTRATO', 'RENTA SIN IVA 2024', 'RENTA SIN IVA 2025', 'Rovida, S.L.', 'RENTA', 'FIANZA', 'ACTUALIZACION RENTA'):
                tenants.append({'name': inquilino, 'rent': renta})
        
        buildings.append({
            'name': sheet_name,
            'address': info['address'],
            'city': info['city'],
            'type': info['type'],
            'tenants': tenants,
            'total_units': max(len(tenants), ws.max_row - 2),
        })
    
    return buildings

def parse_viroda_inmuebles():
    wb = openpyxl.load_workbook('/tmp/INMUEBLES_VIRODA.xlsx')
    buildings = []
    
    sheet_info = {
        'MANUEL SILVELA': {'address': 'C/Manuel Silvela 5, Madrid', 'city': 'Madrid'},
        'HERNANDEZ DE TEJADA': {'address': 'C/Hernández de Tejada 6, Madrid', 'city': 'Madrid'},
        'CANDELARIA MORA': {'address': 'C/Candelaria Mora, Madrid', 'city': 'Madrid'},
        'REINA': {'address': 'C/Reina 15, Madrid', 'city': 'Madrid'},
        'M.PELAYO': {'address': 'C/M. Pelayo 15, Palencia', 'city': 'Palencia'},
    }
    
    for sheet_name in wb.sheetnames:
        info = sheet_info.get(sheet_name, {'address': sheet_name, 'city': 'Madrid'})
        ws = wb[sheet_name]
        tenants = []
        
        for row in range(2, min(ws.max_row + 1, 100)):
            inquilino = None
            renta = None
            finca = None
            tipo = None
            
            for col in range(1, min(ws.max_column + 1, 10)):
                val = ws.cell(row=row, column=col).value
                header = ws.cell(row=1, column=col).value
                
                if header and 'INQUILINO' in str(header).upper() and val:
                    inquilino = str(val).strip()
                elif header and 'RENTA' in str(header).upper() and val:
                    try:
                        renta = float(str(val).replace(',', '.').replace('\xa0', ''))
                    except:
                        pass
                elif header and 'FINCA' in str(header).upper() and val:
                    finca = str(val).strip()
                elif header and 'TIPO' in str(header).upper() and val:
                    tipo = str(val).strip()
            
            if inquilino and inquilino not in ('', 'VACIO', 'USO FAMILIAR') and len(inquilino) > 2:
                tenants.append({
                    'name': inquilino,
                    'rent': renta,
                    'unit': finca,
                    'type': tipo,
                })
        
        buildings.append({
            'name': sheet_name,
            'address': info['address'],
            'city': info['city'],
            'type': 'vivienda',
            'tenants': tenants,
            'total_units': max(len(tenants), 5),
        })
    
    return buildings

def parse_clientes(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    clients = []
    
    for row in range(2, ws.max_row + 1):
        nif = ws.cell(row=row, column=2).value
        razon = ws.cell(row=row, column=3).value
        nombre = ws.cell(row=row, column=4).value
        email = ws.cell(row=row, column=5).value
        telefono = ws.cell(row=row, column=6).value
        activo = ws.cell(row=row, column=9).value
        
        if razon:
            clients.append({
                'dni': str(nif).strip() if nif else '',
                'name': str(razon).strip(),
                'commercial_name': str(nombre).strip() if nombre else '',
                'email': str(email).strip() if email else '',
                'phone': str(telefono).strip() if telefono else '',
                'active': str(activo).strip().lower() == 'sí' if activo else True,
            })
    
    return clients

def main():
    print("Parsing Excel files...")
    rovida_buildings = parse_rovida_inmuebles()
    viroda_buildings = parse_viroda_inmuebles()
    rovida_clients = parse_clientes('/tmp/DATOS_CLIENTES_ROVIDA.xlsx')
    viroda_clients = parse_clientes('/tmp/DATOS_CLIENTES_VIRODA.xlsx')
    
    print(f"Rovida: {len(rovida_buildings)} buildings, {sum(len(b['tenants']) for b in rovida_buildings)} tenants, {len(rovida_clients)} clients")
    print(f"Viroda: {len(viroda_buildings)} buildings, {sum(len(b['tenants']) for b in viroda_buildings)} tenants, {len(viroda_clients)} clients")
    
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(SERVER, username=USER, password=PASSWORD, timeout=15)
    
    # Get company IDs
    out, _ = exec_node(client,
        "require('dotenv').config();require('dotenv').config({path:'.env.local'});"
        "const{PrismaClient}=require('@prisma/client');const p=new PrismaClient();"
        "p.company.findMany({select:{id:true,nombre:true}}).then(c=>{"
        "console.log(JSON.stringify(c));p.$disconnect()}).catch(e=>{console.log(e.message);p.$disconnect()})"
    )
    
    companies = json.loads(out) if out else []
    rovida_id = next((c['id'] for c in companies if 'rovida' in c['nombre'].lower()), None)
    viroda_id = next((c['id'] for c in companies if 'viroda' in c['nombre'].lower()), None)
    
    if not rovida_id:
        rovida_id = 'cef19f55f7b6ce0637d5ffb53'
    if not viroda_id:
        viroda_id = 'cmkctneuh0001nokn7nvhuweq'
    
    print(f"\nRovida company ID: {rovida_id}")
    print(f"Viroda company ID: {viroda_id}")
    
    if not rovida_id or not viroda_id:
        print("ERROR: Companies not found in DB")
        client.close()
        return
    
    # Import tenants for each company
    for company_id, company_name, buildings, clients in [
        (rovida_id, 'Rovida', rovida_buildings, rovida_clients),
        (viroda_id, 'Viroda', viroda_buildings, viroda_clients),
    ]:
        print(f"\n{'='*50}")
        print(f"Importing {company_name} data...")
        
        # Upsert tenants from client list
        imported = 0
        for cl in clients:
            if not cl['name'] or cl['name'] == 'None':
                continue
            
            name_parts = cl['name'].split(',', 1) if ',' in cl['name'] else cl['name'].rsplit(' ', 1)
            apellidos = name_parts[0].strip() if len(name_parts) > 1 else ''
            nombre = name_parts[1].strip() if len(name_parts) > 1 else cl['name'].strip()
            
            dni_escaped = cl['dni'].replace("'", "\\'")
            nombre_escaped = nombre.replace("'", "\\'")
            apellidos_escaped = apellidos.replace("'", "\\'")
            email_escaped = (cl['email'] or '').replace("'", "\\'")
            phone_escaped = (cl['phone'] or '').replace("'", "\\'")
            full_name = f"{nombre} {apellidos}".strip().replace("'", "\\'")
            
            script = (
                "require('dotenv').config();require('dotenv').config({path:'.env.local'});"
                "const{PrismaClient}=require('@prisma/client');const p=new PrismaClient();"
                f"p.tenant.upsert({{where:{{companyId_email:{{companyId:'{company_id}',email:'{email_escaped or full_name.lower().replace(' ','.')+'@placeholder.local'}'}}}},"
                f"update:{{nombreCompleto:'{full_name}',telefono:'{phone_escaped}'}},"
                f"create:{{companyId:'{company_id}',nombre:'{nombre_escaped}',apellidos:'{apellidos_escaped}',"
                f"nombreCompleto:'{full_name}',dni:'{dni_escaped}',"
                f"email:'{email_escaped or full_name.lower().replace(' ','.')+'@placeholder.local'}',"
                f"telefono:'{phone_escaped}',activo:true}}}}).then(t=>{{console.log('OK');p.$disconnect()}}).catch(e=>{{console.log('ERR:'+e.message.substring(0,80));p.$disconnect()}})"
            )
            
            out, _ = exec_node(client, script, timeout=10)
            if 'OK' in (out or ''):
                imported += 1
            
            if imported % 20 == 0 and imported > 0:
                print(f"  {company_name}: {imported} tenants imported...")
        
        print(f"  {company_name}: {imported} tenants imported total")
    
    client.close()
    print("\nImport completed!")

if __name__ == '__main__':
    main()
