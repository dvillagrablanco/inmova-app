#!/usr/bin/env python3
"""
Parser de archivos Excel de clientes y facturación de Rovida y Viroda.

Descarga los archivos de Google Drive y genera parsed-data.json que es
consumido por el script TypeScript import-rovida-viroda-clients.ts

Uso:
  pip install gdown openpyxl
  python3 scripts/parse-rovida-viroda-excel.py

Fuente: https://drive.google.com/drive/folders/1GRgsvuUBV5swHmgB2pvMKWjgQSN9i08h
"""

import os
import sys
import json

# Check dependencies
try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

try:
    import gdown
except ImportError:
    print("Instalando gdown...")
    os.system(f"{sys.executable} -m pip install gdown")
    import gdown

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(PROJECT_DIR, "data-import", "CLIENTES")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "data-import", "parsed-data.json")

GDRIVE_FOLDER = "https://drive.google.com/drive/folders/1GRgsvuUBV5swHmgB2pvMKWjgQSN9i08h"


def download_files():
    """Download files from Google Drive if not present."""
    if os.path.exists(DATA_DIR) and len(os.listdir(DATA_DIR)) >= 4:
        print(f"Archivos ya descargados en {DATA_DIR}")
        return

    os.makedirs(os.path.dirname(DATA_DIR), exist_ok=True)
    print(f"Descargando archivos de Google Drive...")
    gdown.download_folder(
        GDRIVE_FOLDER,
        output=os.path.dirname(DATA_DIR),
        quiet=False,
    )
    print("Descarga completada.")


def parse_clients(filepath, company_name):
    """Parse DATOS CLIENTES file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    clients = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            continue  # Skip header

        nif = str(row[1]).strip() if row[1] else None
        razon_social = str(row[2]).strip() if row[2] else None
        nombre_comercial = str(row[3]).strip() if row[3] else None
        email = str(row[4]).strip() if row[4] else None
        telefono = str(row[5]).strip() if row[5] else None
        activo = str(row[8]).strip() if row[8] else "No"
        medio_pago = str(row[9]).strip() if row[9] else None
        iban = str(row[11]).strip() if row[11] else None

        tipo_via = str(row[15]).strip() if row[15] else ""
        direccion = str(row[16]).strip() if row[16] else ""
        numero = str(row[17]).strip() if row[17] else ""
        piso = str(row[18]).strip() if row[18] else ""
        escalera = str(row[19]).strip() if row[19] else ""
        puerta = str(row[20]).strip() if row[20] else ""
        cp = str(row[21]).strip() if row[21] else ""
        poblacion = str(row[22]).strip() if row[22] else ""
        provincia = str(row[23]).strip() if row[23] else ""
        pais = str(row[24]).strip() if row[24] else "ESPAÑA"

        if not nif or nif == "None":
            continue

        # Build full address
        addr_parts = []
        for val, prefix in [
            (tipo_via, ""), (direccion, ""), (numero, ""),
            (piso, "Piso "), (escalera, "Esc. "), (puerta, "Pta. "),
            (cp, ""), (poblacion, ""),
        ]:
            if val and val != "None":
                addr_parts.append(f"{prefix}{val}" if prefix else val)
        if provincia and provincia != "None" and provincia != poblacion:
            addr_parts.append(provincia)

        nombre = nombre_comercial if nombre_comercial and nombre_comercial != "None" else razon_social

        clients.append({
            "nif": nif,
            "nombreCompleto": nombre,
            "email": email if email and email != "None" else None,
            "telefono": telefono if telefono and telefono != "None" else None,
            "activo": activo == "Sí",
            "medioPago": medio_pago if medio_pago and medio_pago != "None" else None,
            "iban": iban if iban and iban != "None" else None,
            "direccion": ", ".join(addr_parts) if addr_parts else None,
            "poblacion": poblacion if poblacion and poblacion != "None" else None,
            "provincia": provincia if provincia and provincia != "None" else None,
            "codigoPostal": cp if cp and cp != "None" else None,
            "pais": pais if pais and pais != "None" else "ESPAÑA",
            "company": company_name,
        })

    return clients


def parse_invoices(filepath, company_name):
    """Parse FACTURACION file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    invoices = []
    current_invoice = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= 2:
            continue

        factura_num = row[0]
        nif = str(row[3]).strip() if row[3] else None
        nombre = str(row[4]).strip() if row[4] else None
        operacion = str(row[6]).strip() if row[6] else None
        base_importe = row[7]
        total = row[10]
        concepto = str(row[11]).strip() if row[11] else None
        precio_ud = row[13]

        if factura_num and factura_num != "None":
            current_invoice = {
                "numFactura": str(factura_num),
                "fecha": str(row[2]) if row[2] else None,
                "nif": nif,
                "nombre": nombre,
                "operacion": operacion,
                "baseImporte": float(base_importe) if base_importe else 0,
                "total": float(total) if total else 0,
                "lineas": [],
                "company": company_name,
            }
            invoices.append(current_invoice)

        if current_invoice and concepto and concepto != "None":
            precio = float(precio_ud) if precio_ud else 0
            is_rent = "Renta" in concepto or "Alquiler" in concepto
            current_invoice["lineas"].append({
                "concepto": concepto,
                "precio": precio,
                "isRent": is_rent,
            })

    return invoices


def extract_rent_mappings(invoices):
    """Extract rent info: NIF -> [{building, unit, rent}]."""
    mappings = {}
    for inv in invoices:
        nif = inv["nif"]
        if not nif:
            continue

        rent_total = 0
        unit_info = None
        for linea in inv["lineas"]:
            if linea["isRent"]:
                rent_total += linea["precio"]
                unit_info = linea["concepto"]

        if rent_total > 0:
            if nif not in mappings:
                mappings[nif] = []
            mappings[nif].append({
                "nombre": inv["nombre"],
                "rentaMensual": rent_total,
                "operacion": inv.get("operacion", ""),
                "concepto": unit_info,
                "totalFactura": inv["total"],
                "numFactura": inv["numFactura"],
            })

    return mappings


def main():
    print("=" * 70)
    print("  PARSER: Datos clientes Rovida y Viroda")
    print("=" * 70)

    # Download
    download_files()

    # Parse
    rovida_path = os.path.join(DATA_DIR, "DATOS CLIENTES ROVIDA.xlsx")
    viroda_path = os.path.join(DATA_DIR, "DATOS CLIENTES VIRODA.xlsx")
    rovida_fact = os.path.join(DATA_DIR, "FACTURACION ROVIDA FEB.xlsx")
    viroda_fact = os.path.join(DATA_DIR, "FACTURACION VIRODA FEB.xlsx")

    print("\nParseando DATOS CLIENTES ROVIDA...")
    rovida_clients = parse_clients(rovida_path, "ROVIDA")
    print(f"  -> {len(rovida_clients)} clientes")

    print("Parseando DATOS CLIENTES VIRODA...")
    viroda_clients = parse_clients(viroda_path, "VIRODA")
    print(f"  -> {len(viroda_clients)} clientes")

    print("Parseando FACTURACION ROVIDA FEB...")
    rovida_invoices = parse_invoices(rovida_fact, "ROVIDA")
    print(f"  -> {len(rovida_invoices)} facturas")

    print("Parseando FACTURACION VIRODA FEB...")
    viroda_invoices = parse_invoices(viroda_fact, "VIRODA")
    print(f"  -> {len(viroda_invoices)} facturas")

    rovida_rents = extract_rent_mappings(rovida_invoices)
    viroda_rents = extract_rent_mappings(viroda_invoices)

    output = {
        "rovida": {
            "clients": rovida_clients,
            "invoices": rovida_invoices,
            "rentMappings": rovida_rents,
        },
        "viroda": {
            "clients": viroda_clients,
            "invoices": viroda_invoices,
            "rentMappings": viroda_rents,
        },
    }

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nRent mappings ROVIDA: {len(rovida_rents)} inquilinos con renta")
    print(f"Rent mappings VIRODA: {len(viroda_rents)} inquilinos con renta")
    print(f"\nDatos guardados en: {OUTPUT_FILE}")
    print("\nSiguiente paso: npx tsx scripts/import-rovida-viroda-clients.ts")


if __name__ == "__main__":
    main()
